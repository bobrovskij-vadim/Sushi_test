import time
import os.path
import openpyxl
import read_from_file
import business_logic

class Write_to_file:
    def __init__(self):
        self.__date = time.strftime("%Y-%m", time.localtime())
        extension = ".xlsx"
        self.__file_name = self.__date + extension

    def add_to_month_result(self, position_list):
        if os.path.isfile(self.__file_name):
            self.__book = openpyxl.open(self.__file_name)
            self.__sheet = self.__book.active
            self.__ex_pos_list = read_from_file.Read_from_file(self.__file_name).get_position_list()
            for i in range(0, len(self.__ex_pos_list)):
                self.__ex_pos_list[i].set_amount(int(position_list[i].get_amount()))
            self.save_to_file(self.__ex_pos_list, self.__sheet)
            self.__book.save(self.__file_name)
            self.__book.close()
        else:
            work_book = openpyxl.Workbook()
            work_sheet = work_book.active
            work_sheet.title = self.__date
            work_sheet.cell(row=1, column=1).value = "Ном поз"
            work_sheet.cell(row=1, column=2).value = "Наименование"
            work_sheet.cell(row=1, column=3).value = "Цена"
            work_sheet.cell(row=1, column=4).value = "Скидка"
            work_sheet.cell(row=1, column=5).value = "Количество"
            #work_sheet.cell(row=1, column=6).value = "Сумма"
            self.save_to_file(position_list, work_sheet)
            work_book.save(self.__file_name)
            work_book.close()

    def save_changed_order(self, position, amount):
        self.__book = openpyxl.open(self.__file_name)
        self.__sheet = self.__book.active
        self.__ex_pos_list = read_from_file.Read_from_file(self.__file_name).get_position_list()

        for i in range(0, len(self.__ex_pos_list)):
            if self.__ex_pos_list[i].get_position() == position.get_position():
                self.__ex_pos_list[i].set_amount(int(amount))

        self.save_to_file(self.__ex_pos_list, self.__sheet)
        self.__book.save(self.__file_name)
        self.__book.close()

    def save_to_file(self, position_list, work_sheet):
        r = 2
        for i in position_list:
            work_sheet.cell(row=r, column=1).value = i.get_position()
            work_sheet.cell(row=r, column=2).value = i.get_name()
            work_sheet.cell(row=r, column=3).value = i.get_price()
            work_sheet.cell(row=r, column=4).value = i.get_discount()
            work_sheet.cell(row=r, column=5).value = i.get_amount()
            r += 1





